from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect
from django.urls import reverse
from django.db.models import Max
from django.utils.html import strip_tags
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Laboratorio, Infraestrutura,LaboratorioInfraestrutura,ImagemLaboratorio, Marca, Equipamento, RegimentoInterno, UnidadeAcademica,ImagemInfraestrutura,GrupoDePesquisa,MembroLaboratorio
from django.contrib.messages import constants
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from PIL import Image
from datetime import datetime
import io
import base64
from django.core.files.base import ContentFile
import logging
from django.templatetags.static import static
from django.core.exceptions import ValidationError
from django.http import FileResponse
from django.core.files import File
from .Send import Email
from .planilha import export_to_excel

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
import pandas as pd
from openpyxl.styles import PatternFill
from .models import Unidade, Laboratorio, Equipamento, RegimentoInterno, Unidade, Apresentacao, Pos, Formento
 # Adicione outros modelos conforme necessário
from django.db.models import Q, Max
from django.core.exceptions import PermissionDenied



from django.db import connections
from .models import Projeto
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import Projeto, Laboratorio
from .forms import ProjetoForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from math import isnan
from django.contrib.auth.decorators import user_passes_test

from django.http import HttpResponseForbidden
from .forms import ExclusaoRegimentoInternoForm, ExclusaounidadeForm


def requer_padace(padace_id):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                raise PermissionDenied("Você precisa estar autenticado para executar esta ação.")

            try:
                # Busca o PADACE do usuário autenticado no modelo UsuarioInfo
                usuario_info = UsuarioInfo.objects.get(user=request.user)
                if usuario_info.padaces != padace_id:
                    raise PermissionDenied("Você não tem permissão para executar esta ação.")
            except UsuarioInfo.DoesNotExist:
                raise PermissionDenied("Informações adicionais do usuário não encontradas.")

            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator
@csrf_exempt
def visualizar_pdf(request, unidade_academica_id):
    try:
        unidade_academica = UnidadeAcademica.objects.get(id=unidade_academica_id)
        
        if unidade_academica and unidade_academica.pdf:
            pdf_file = unidade_academica.pdf
            response = FileResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{pdf_file.name}"'
            return response
    except UnidadeAcademica.DoesNotExist:
        pass

    return HttpResponse("PDF não encontradoooo.", status=404)

def visualizar_regimento_interno(request, regimento_id):
    try:
        regimento = RegimentoInterno.objects.get(id=regimento_id)

        if regimento and regimento.pdf:
            pdf_file = regimento.pdf
            response = FileResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{pdf_file.name}"'
            return response
    except RegimentoInterno.DoesNotExist:
        pass

    return HttpResponse("Regimento Interno não encontrado.", status=404)

@csrf_exempt
def editar_regimentos_internos(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    regimentos_internos = RegimentoInterno.objects.filter(laboratorio=laboratorio)
    unidades_academicas = UnidadeAcademica.objects.filter(laboratorio=laboratorio)

    # Verifica se o usuário é superusuário
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        regimento_pdf = request.FILES.get('regimento_pdf')
        unidade_academica_pdf = request.FILES.get('unidade_academica_pdf')

        # Se o usuário não for admin, retorne uma mensagem de erro ou redirecione conforme necessário
        if not is_admin:
            return HttpResponse("Você não tem permissão para adicionar.", status=403)

        if regimento_pdf:
            regimento_interno = RegimentoInterno(
                laboratorio=laboratorio,
                pdf=regimento_pdf
            )
            regimento_interno.save()
            return redirect('editar_regimentos_internos', laboratorio_id=laboratorio_id)
        
        if unidade_academica_pdf:
            # Verifica se já existe uma Unidade Acadêmica associada a este laboratório
            unidade_academica_existente = UnidadeAcademica.objects.filter(laboratorio=laboratorio).first()

            # Se existir, substitui o PDF existente
            if unidade_academica_existente:
                unidade_academica_existente.pdf = unidade_academica_pdf
                unidade_academica_existente.save()
            else:
                # Se não existir, cria uma nova Unidade Acadêmica
                UnidadeAcademica.objects.create(laboratorio=laboratorio, pdf=unidade_academica_pdf)

            return redirect('editar_regimentos_internos', laboratorio_id=laboratorio_id)

    return render(request, 'editar_regimentos_internos.html', {
        'laboratorio': laboratorio,
        'regimentos_internos': regimentos_internos,
        'unidades_academicas': unidades_academicas,
        'is_admin': is_admin,
    })



@csrf_exempt
def excluir_unidade(request, unidade_id):
    unidade = get_object_or_404(UnidadeAcademica, id=unidade_id)

    if request.method == 'POST':
        form = ExclusaounidadeForm(request.POST)

        if form.is_valid():
            unidade.delete()
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@csrf_exempt
def excluir_regimento_interno(request, regimento_id):
    regimento = get_object_or_404(RegimentoInterno, id=regimento_id)

    if request.method == 'POST':
        form = ExclusaoRegimentoInternoForm(request.POST)

        if form.is_valid():
            regimento.delete()
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return redirect('editar_regimentos_internos', laboratorio_id=regimento.laboratorio.id)


@csrf_exempt
@requer_padace(708)
def excluir_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, pk=laboratorio_id)
    laboratorio.delete()
    messages.success(request, 'Laboratório excluído com sucesso!')
    return redirect('main')

@csrf_exempt
def editar_unidades_academicas(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    unidades_academicas = UnidadeAcademica.objects.filter(laboratorio=laboratorio)

    # Verifica se o usuário é superusuário
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        if not is_admin:
            return HttpResponse("Você não tem permissão para adicionar unidades acadêmicas.", status=403)

        unidade_academica_pdf = request.FILES.get('unidade_academica_pdf')
        if unidade_academica_pdf:
            UnidadeAcademica.objects.create(laboratorio=laboratorio, pdf=unidade_academica_pdf)

    return render(request, 'editar_unidades_academicas.html', {
        'laboratorio': laboratorio,
        'unidades_academicas': unidades_academicas,
        'is_admin': is_admin,
    })

@csrf_exempt
def visualizar_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    user_ldap = request.user.username if request.user.is_authenticated else ''
    is_admin = request.user.is_superuser

    # Verifica se o usuário é o responsável ou associado ao laboratório
    tem_permissao = (
        is_admin or
        laboratorio.user_ldap_responsavel.lower() == user_ldap.lower() or
        laboratorio.responsaveis_associados.filter(user_ldap__iexact=user_ldap).exists()
    )

    if not tem_permissao:
        return render(request, 'acesso_restrito.html', {
            'mensagem': 'Você não tem permissão para acessar este laboratório.',
            'laboratorio': laboratorio,
        })

    regimentos_internos = RegimentoInterno.objects.filter(laboratorio=laboratorio)
    unidades_academicas = UnidadeAcademica.objects.filter(laboratorio=laboratorio)
    infraestruturas = Infraestrutura.objects.filter(laboratorio_id=laboratorio.id, status=True)
    projetos = Projeto.objects.filter(laboratorio=laboratorio)
    membros = MembroLaboratorio.objects.filter(laboratorio=laboratorio)
    grupos_de_pesquisa = laboratorio.grupos_de_pesquisa.all()
    programas = Pos.objects.filter(laboratorio=laboratorio_id)
    formento = Formento.objects.filter(laboratorio=laboratorio_id)

    return render(request, 'view.html', {
        'laboratorio': laboratorio,
        'infraestruturas': infraestruturas,
        'regimentos_internos': regimentos_internos,
        'unidades_academicas': unidades_academicas,
        'projetos': projetos,
        'membros': membros,
        'grupos_de_pesquisa': grupos_de_pesquisa,
        'data_atualizacao': laboratorio.updated_at,
        'programas': programas,
        'formento': formento,
        'is_admin': is_admin,
    })

@csrf_exempt
def index(request):
    return render(request, 'index.html')

@csrf_exempt
def redirecionar_para_geral(request):
    return redirect('geral')

@login_required
@csrf_exempt
def edit(request):
   
    laboratorios = Laboratorio.objects.all()
    marcas = Marca.objects.all()
    equipamentos = Equipamento.objects.all()
    formento = Formento.objects.all()

    return render(request, 'edit.html', {
        'laboratorios': laboratorios,
        'marcas': marcas,
        'equipamentos': equipamentos,
        'formento': formento,
    })

@csrf_exempt
def imagem_rgb(request, imagem_id):
    try:
        imagem = ImagemLaboratorio.objects.get(pk=imagem_id)
        imagem_bin = imagem.imagem

        response = HttpResponse(content_type="image/jpeg")
        image = Image.open(io.BytesIO(imagem_bin))
        
        # Redimensionar a imagem para o tamanho desejado (600x900)
        # new_size = (600, 900)
        # image = image.resize(new_size, Image.ANTIALIAS)
        
        image = image.convert("RGB")
        image.save(response, format="JPEG")

        return response
    except ImagemLaboratorio.DoesNotExist:
        return HttpResponse("Imagem não encontrada.", status=404)

    
@login_required 
@csrf_exempt
def main(request):
    texto = request.POST.get('filtro', '')
    nome_laboratorio = request.POST.get('nome_laboratorio', '')
    nome_do_grupo = request.POST.get('grupo', '')
    responsavel = request.POST.get('responsavel', '')
    unidade = request.POST.get('unidade', '')
    pos = request.POST.get('pos', '')
    fomento = request.POST.get('fomento', '')
    projeto = request.POST.get('projeto', '')
    equipamento = request.POST.get('equipamento', '')
    aprovado = request.POST.get('aprovado', '')

    user_ldap_session = request.user.username if request.user.is_authenticated else ''

    laboratorios = Laboratorio.objects.all()

    # Se o usuário NÃO for superuser, só vê laboratórios em que é responsável
    if not request.user.is_superuser:
        laboratorios = laboratorios.filter(
            Q(user_ldap_responsavel__iexact=user_ldap_session) |
            Q(responsaveis_associados__user_ldap__iexact=user_ldap_session)
        ).distinct()

    conditions = Q()
    if texto:
        laboratorios = laboratorios.filter(nome_laboratorio__icontains=texto)
    else:
        if nome_laboratorio:
            conditions &= Q(nome_laboratorio__icontains=nome_laboratorio)
        if responsavel:
            conditions &= Q(responsavel__icontains=responsavel)
        if unidade:
            conditions &= Q(unidade__icontains=unidade)
        if pos:
            conditions &= Q(pos__nome_do_Programa__icontains=pos)
        if fomento:
            conditions &= Q(Formentos__numerodoformento__icontains=fomento)
        if nome_do_grupo:
            conditions &= Q(grupos_de_pesquisa__nome_do_grupo__icontains=nome_do_grupo)
        if equipamento:
            conditions &= Q(infraestrutura__equipamento__nome_equipamento__icontains=equipamento)
        if aprovado:
            conditions &= Q(aprovado=aprovado)

        laboratorios = laboratorios.filter(conditions)

    paginator = Paginator(laboratorios, 10)
    page = request.GET.get('page')
    laboratorios = paginator.get_page(page)

    lab_count = Laboratorio.objects.count()
    infra_count = Equipamento.objects.count()
    texto = Apresentacao.objects.get(pk=1)
    options = Equipamento.objects.filter(nome_equipamento__icontains=equipamento)

    context = {
        'laboratorios': laboratorios,
        'is_admin': request.user.is_superuser,
        'lab_count': lab_count,
        'infra_count': infra_count,
        'user_ldap_responsavel': user_ldap_session,
        'texto': texto,
        'options': options,
        'num': laboratorios.paginator.count,
    }

    return render(request, 'main.html', context)

@login_required
@csrf_exempt
def texto_ajustavel(request):
    if request.method == "POST":
        texto_ajustavel = request.POST.get('texto_ajustavel', '')

        texto = Apresentacao.objects.get(pk=1)
        texto.texto_ajustavel = texto_ajustavel
        texto.save()

    return redirect('main')
    
 

@login_required
@csrf_exempt
@user_passes_test(lambda u: u.is_superuser)
def salvar_laboratorio(request):
    try:
        if request.method == "POST":
            imagens_lab = request.FILES.getlist("imagens_lab[]")
            imagens_salvas = []

            for imagem in imagens_lab:
                imagem_laboratorio = ImagemLaboratorio(imagem=imagem)
                imagem_laboratorio.save()
                imagens_salvas.append(imagem_laboratorio)

            if not imagens_lab:
                caminho_imagem_padrao = 'templates/static/images/microgrande.png'
                with open(caminho_imagem_padrao, 'rb') as img_padrao:
                    imagem_laboratorio = ImagemLaboratorio(imagem=File(img_padrao))
                    imagem_laboratorio.save()
                    imagens_salvas.append(imagem_laboratorio)

            nome_laboratorio = request.POST.get("nome_laboratorio")
            responsavel = request.POST.get("responsavel")
            cpf_responsavel = request.POST.get('cpf_responsavel')
            user_ldap_responsavel = request.POST.get('user_ldap_responsavel')
            email = request.POST.get("email")
            telefone = request.POST.get("telefone")
            unidade = request.POST.get("unidade")
            rua = request.POST.get("rua")
            numero_rua = request.POST.get("numero_rua")
            cep = request.POST.get("cep")
            bairro = request.POST.get("bairro")
            ato_anexo = request.FILES.get("ato_anexo")
            ato_anexo_content = ato_anexo.read() if ato_anexo else None
            descricao = request.POST.get("descricao") or None
            link_pnipe = request.POST.get("link_pnipe") or None
            andar = request.POST.get("andar") or None
            sala = request.POST.get("sala") or None
            unidade = unidade or None
            bairro = bairro or None
            rua = rua or None
            numero_rua = numero_rua or None
            cep = cep or None

            equipamento_id = request.POST.get('equipamento')
            marca_id = request.POST.get('marca')
            modelo = request.POST.get('modelo') or None
            finalidade = request.POST.get('finalidade') or None

            try:
                equipamento = Equipamento.objects.get(id=equipamento_id)
            except Equipamento.DoesNotExist:
                equipamento = None

            try:
                marca = Marca.objects.get(id=marca_id)
            except Marca.DoesNotExist:
                marca = None

            # Enviar e-mail (função definida em outro lugar)
            Email(email, nome_laboratorio, responsavel)

            laboratorio = Laboratorio.objects.create(
                nome_laboratorio=nome_laboratorio,
                responsavel=responsavel,
                email=email,
                telefone=telefone,
                unidade=unidade,
                rua=rua,
                numero_rua=numero_rua,
                cep=cep,
                bairro=bairro,
                ato_anexo=ato_anexo_content,
                descricao=descricao,
                link_pnipe=link_pnipe,
                cpf_responsavel=cpf_responsavel,
                user_ldap_responsavel=user_ldap_responsavel,
            )

            if request.FILES.get("pdf_unidade_academica"):
                pdf = request.FILES["pdf_unidade_academica"]
                if not pdf.name.endswith(".pdf"):
                    raise ValidationError("O arquivo deve ser um PDF.")
                UnidadeAcademica.objects.create(laboratorio=laboratorio, pdf=pdf)

            infraestrutura = Infraestrutura.objects.create(
                equipamento=equipamento,
                marca=marca,
                modelo=modelo,
                finalidade=finalidade,
            )

            laboratorio.infraestrutura = infraestrutura
            laboratorio.save()
            laboratorio.imagens.set(imagens_salvas)

            if imagens_salvas:
                return HttpResponseRedirect(reverse('editar_laboratorio', args=[laboratorio.id, imagens_salvas[0].id]))
            else:
                messages.add_message(request, messages.WARNING, 'Nenhuma imagem foi salva.')
                return HttpResponseRedirect(reverse('editar_laboratorio', args=[laboratorio.id, 0]))
    except Exception as e:
        messages.add_message(request, messages.WARNING, 'O sistema apresenta falhas internas.')

    return redirect('../edit/')

from django.contrib.auth.models import User
@login_required 
@csrf_exempt
def buscar_nomes(request):
    term = request.GET.get('term', '')
    if term:
        usuarios = User.objects.filter(
            is_superuser=False,
            first_name__icontains=term
        )[:10]

        results = []
        for usuario in usuarios:
            results.append({
                'label': usuario.get_full_name() or usuario.username,
                'cpf': '',  # Preencha aqui se você tiver o CPF em outro modelo
                'user_ldap': usuario.username,
            })
        return JsonResponse(results, safe=False)

    return JsonResponse([], safe=False)


@login_required
@csrf_exempt
def adicionar_grupo_de_pesquisa(request, laboratorio_id=None):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else request.user.laboratorio

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        # Restante do código para processar os campos do formulário e salvar as alterações
        nome_do_grupo = request.POST.get('nome_do_grupo')
        area = request.POST.get('area')
        link_grupo = request.POST.get('link_grupo')

        # Crie uma nova instância de GrupoDePesquisa com os dados fornecidos
        grupo_de_pesquisa = GrupoDePesquisa.objects.create(
            nome_do_grupo=nome_do_grupo,
            area=area,
            link_grupo=link_grupo
        )

        # Associe o grupo de pesquisa ao laboratório atual
        laboratorio.grupos_de_pesquisa.add(grupo_de_pesquisa)

        # Redirecione para a página de edição do laboratório ou para onde for adequado
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    grupos_de_pesquisa = laboratorio.grupos_de_pesquisa.all()

    return render(
        request,
        'adicionar_grupo_de_pesquisa.html',
        {'laboratorio': laboratorio, 'grupos_de_pesquisa': grupos_de_pesquisa, 'is_admin': is_admin}
    )

@login_required
@csrf_exempt
def editar_grupo_de_pesquisa(request, grupo_de_pesquisa_id):
    grupo_de_pesquisa = get_object_or_404(GrupoDePesquisa, id=grupo_de_pesquisa_id)

    if request.method == 'POST':
        grupo_de_pesquisa_id = request.POST.get('grupo_de_pesquisa_id')  # Recupere o ID do grupo de pesquisa
        nome_do_grupo = request.POST.get('nome_do_grupo')
        area = request.POST.get('area')
        link_grupo = request.POST.get('link_grupo')

        # Verifique se o ID no formulário corresponde ao ID do grupo de pesquisa
        if grupo_de_pesquisa_id == str(grupo_de_pesquisa.id):
            # Atualize os valores do grupo de pesquisa
            grupo_de_pesquisa.nome_do_grupo = nome_do_grupo
            grupo_de_pesquisa.area = area
            grupo_de_pesquisa.link_grupo = link_grupo
            grupo_de_pesquisa.save()

            # Redirecione de volta para a página original ou para onde for apropriado
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # Renderize a página de edição do grupo de pesquisa
    return render(request, 'editar_grupo_de_pesquisa.html', {'grupo_de_pesquisa': grupo_de_pesquisa})


# novo
@login_required
@csrf_exempt
def excluir_grupo_de_pesquisa(request, grupo_de_pesquisa_id):
    grupo_de_pesquisa = get_object_or_404(GrupoDePesquisa, id=grupo_de_pesquisa_id)

    grupo_de_pesquisa.delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
@csrf_exempt
def adicionar_equipamento(request):
   
    if request.method == 'POST':
        nome_equipamento = request.POST.get('nome_equipamento')
        nome_marca = request.POST.get('nome_marca')

        if nome_marca:

            marca_existente = Marca.objects.filter(
                nome_marca=nome_marca
            ).exists()

            if not marca_existente:
                marca = Marca.objects.create(
                    nome_marca=nome_marca
 )
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                return HttpResponse('Marca já existe.')

        if nome_equipamento:

            equipamento_existente = Equipamento.objects.filter(
                nome_equipamento=nome_equipamento,
            ).exists()

            if not equipamento_existente:
                equipamento = Equipamento.objects.create(
                    nome_equipamento=nome_equipamento,
                )
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                # Adiciona a mensagem de erro ao sistema de mensagens
                messages.error(request, 'Equipamento já existe.')
   
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
@csrf_exempt
def editar_infraestrutura(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    marcas = Marca.objects.all()
    equipamentos = Equipamento.objects.all()

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        equipamento_id = request.POST.get('equipamento')
        marca_id = request.POST.get('marca')
        modelo = request.POST.get('modelo')
        finalidade = request.POST.get('finalidade')
        tombo = request.POST.get('tombo')
        quantidade = request.POST.get('quantidade')
        # Verifique se uma imagem foi enviada
        novas_imagens = request.FILES.getlist('nova_imagem')

        # Crie uma nova infraestrutura
        infraestrutura = Infraestrutura.objects.create(
            equipamento_id=equipamento_id,
            marca_id=marca_id,
            laboratorio_id=laboratorio_id,
            modelo=modelo,
            finalidade=finalidade,
            tombo=tombo,
            quantidade=quantidade,
        )

        # Se uma nova imagem foi enviada, crie uma instância de ImagemInfraestrutura para cada imagem
        for nova_imagem in novas_imagens:
            imagem_infraestrutura = ImagemInfraestrutura(imagem=nova_imagem, infraestrutura=infraestrutura)
            imagem_infraestrutura.save()

        print('Infraestrutura criada com sucesso!')
        # Redirecione para a página de edição do laboratório ou para onde for adequado
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    

    infraestruturas = Infraestrutura.objects.filter(laboratorio_id=laboratorio_id)
    return render(
        request,
        'editar_infraestrutura.html',
        {'laboratorio': laboratorio, 'marcas': marcas, 'equipamentos': equipamentos, 'infraestruturas': infraestruturas,
         'is_admin': is_admin}
    )
@login_required
@csrf_exempt
def equipamento(request):
   
    laboratorios = Laboratorio.objects.all()
  
    return render(request, 'equipamento.html', {
        'laboratorios': laboratorios,
    })

@csrf_exempt
def salvar_equipamento(request):
    if request.method == "POST":
        try:   
            equipamento = request.POST.get("equipamento")
            marca = request.POST.get("marca")
            modelo = request.POST.get("modelo")
            finalidade = request.POST.get("finalidade")
          
            
  # Salvar os dados no banco de dados
            Infraestrutura.objects.create(
                equipamento=equipamento,
                marca=marca,
                modelo=modelo,
                finalidade=finalidade,
                
            )

            messages.add_message(request, messages.SUCCESS, 'Salvo com sucesso.')
            return render(request, 'equipamento.html')
        except:
            messages.add_message(request, messages.WARNING, 'O sistema apresenta falhas internas.')
            return redirect('../equipamento/')
    else:
        return render(request, 'equipamento.html')

@login_required
@csrf_exempt
def paginar(list, limit_per_page, request): 
    paginator = Paginator(list, limit_per_page) 
    page = request.GET.get('page')
    atos = paginator.get_page(page)
    context = {'atos': atos}
    return context

@login_required
def aprovados(request):
    atos_list = AtoNormativ.objects.filter(status='aprovado')
    context = paginar(atos_list, 16, request);
    return render(request, 'main.html', context)

@login_required
@csrf_exempt
def pesquisar(request):
    if request.method == 'GET':
        termo_pesquisa = request.GET.get('pesquisa', '')  # Obtém o termo de pesquisa do parâmetro GET 'pesquisa'
        atos = AtoNormativ.objects.filter(texto_normativo__icontains=termo_pesquisa)  # Filtra os atos com base no termo de pesquisa
        context = {'atos': atos}
        return render(request, 'resultado_pesquisa.html', context)
    
@login_required
@csrf_exempt
def visualizar_arquivo(request, laboratorio_id):
    try:
        laboratorio = Laboratorio.objects.get(pk=laboratorio_id)
        if laboratorio.ato_anexo:
            response = HttpResponse(laboratorio.ato_anexo, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="arquivo.pdf"'  # Pode ajustar o nome do arquivo
            return response
    except Laboratorio.DoesNotExist:
        pass

    return HttpResponse("Arquivo não encontrado.", status=404)

@login_required
@csrf_exempt
def visualizar_imagens(request, laboratorio_id):
    laboratorio = Laboratorio.objects.get(pk=laboratorio_id)
    imagens = laboratorio.imagens.all()
    return render(request, 'visualizar_imagens.html', {'laboratorio': laboratorio, 'imagens': imagens})

@login_required
@csrf_exempt
def view_imagem(request, imagem_id):
    try:
        imagem = ImagemLaboratorio.objects.get(pk=imagem_id)
        imagem_bin = imagem.imagem

        # Configurar o cabeçalho de tipo de conteúdo para uma imagem
        response = HttpResponse(content_type="image/jpeg")
        
        # Abra a imagem usando o Pillow (PIL)
        image = Image.open(io.BytesIO(imagem_bin))

        # Converta a imagem para o modo RGB
        image = image.convert("RGB")
        
        # Salve a imagem no formato JPEG
        image.save(response, format="JPEG")

        return response
    except ObjectDoesNotExist:
        return HttpResponse("Imagem não encontrada.", status=404)
    





from .models import ResponsavelAssociado
import os
import subprocess
import psutil
from django.core.wsgi import get_wsgi_application





import psutil

def restart_gunicorn(request):
    # Encontrar o processo do Gunicorn
    
    for proc in psutil.process_iter(['pid', 'name']):
        if 'gunicorn' in proc.info['name']:
            gunicorn_pid = proc.pid
            break
    else:
        logging.error("Gunicorn process not found")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # Reiniciar o processo do Gunicorn enviando um sinal HUP
    try:
        gunicorn_process = psutil.Process(gunicorn_pid)
        gunicorn_process.send_signal(psutil.signal.SIGHUP)
        logging.info("Gunicorn restarted successfully")
    except Exception as e:
        logging.error(f"Failed to restart Gunicorn: {str(e)}")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
@csrf_exempt
def editar_laboratorio(request, laboratorio_id, imagem_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        # Processar os campos do formulário
        laboratorio.nome_laboratorio = request.POST.get('nome_laboratorio', '')
        laboratorio.responsavel = request.POST.get('responsavel', '')
        laboratorio.email = request.POST.get('email', '')
        laboratorio.telefone = request.POST.get('telefone', '')
        laboratorio.data_criacao = request.POST.get('data_criacao', '')
        laboratorio.apresentacao = request.POST.get('apresentacao', '')
        laboratorio.objetivos = request.POST.get('objetivos', '')
        laboratorio.descricao = request.POST.get('descricao', '')
        laboratorio.link_pnipe = request.POST.get('link_pnipe', '')
        laboratorio.user_ldap_responsavel = request.POST.get('user_ldap_responsavel', '')
        laboratorio.site = request.POST.get('site', '')

        # Processar os novos responsáveis associados
        novos_responsaveis_associados = request.POST.getlist('novo_responsavel_1', '')

        # Adicione estas variáveis para armazenar os dados do novo responsável associado
        novo_responsavel_cpf = None
        novo_responsavel_user_ldap = None

        for nome_responsavel in novos_responsaveis_associados:
            cpf_responsavel = request.POST.get(f'cpf_responsavel_1', '')
            user_ldap_responsavel = request.POST.get(f'user_ldap_responsavel_1', '')

            responsavel_associado = ResponsavelAssociado.objects.create(
                nome=nome_responsavel,
                cpf=cpf_responsavel,
                user_ldap=user_ldap_responsavel,
                laboratorio=laboratorio
            )
            laboratorio.responsaveis_associados.add(responsavel_associado)

            # Atualize as variáveis com os dados do último responsável associado
            novo_responsavel_cpf = responsavel_associado.cpf
            novo_responsavel_user_ldap = responsavel_associado.user_ldap

        pagina_atual = 'editar_laboratorio'

        # Processar a imagem, se presente
        imagem_laboratorio = request.FILES.get('imagem_laboratorio')
        if imagem_laboratorio:
            # Crie uma instância de ImagemLaboratorio associada ao laboratório
            imagem = ImagemLaboratorio(imagem=imagem_laboratorio)
            imagem.save()
            laboratorio.imagens.add(imagem)

        laboratorio.save()
       
        messages.success(request, 'Laboratório atualizado com sucesso.')
        return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=imagem_id)

    # Restante do código para renderizar a página de edição
    return render(request, 'editar_laboratorio.html', {
        'laboratorio': laboratorio,
        'is_admin': is_admin,  # Verificação de permissão baseada no superuser
    })



@login_required
@csrf_exempt
def update_oculto(request, lab_id):
    check = request.POST.get('teste')
    print(check)
    laboratorio = get_object_or_404(Laboratorio, id=lab_id)

    laboratorio.ocultar_laboratorio = check

    laboratorio.save()

    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)

@csrf_exempt
def excluir_imagem_laboratorio(request, laboratorio_id, imagem_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    imagem = get_object_or_404(ImagemLaboratorio, id=imagem_id)

    # Remova a imagem do laboratório
    laboratorio.imagens.remove(imagem)

    # Exclua a imagem do banco de dados
    imagem.delete()

    # Redirecione de volta à página de edição do laboratório
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from .models import Laboratorio, ResponsavelAssociado
from django.urls import reverse


@csrf_exempt
def excluir_responsavel_associado(request, laboratorio_id, responsavel_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    responsavel_associado = get_object_or_404(ResponsavelAssociado, id=responsavel_id)

    # Certifique-se de que o responsável associado está vinculado ao laboratório antes de excluir
    if responsavel_associado in laboratorio.responsaveis_associados.all():
        responsavel_associado.delete()
        messages.success(request, 'Responsável associado excluído com sucesso.')
    else:
        messages.error(request, 'Erro ao excluir responsável associado.')
        print(f"DEBUG: Responsável associado não encontrado no laboratório {laboratorio.id}")

    # Redirecione de volta para a mesma página
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)








@login_required
@csrf_exempt
def editar_endereco(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, pk=laboratorio_id)

    # Dados mocados para substituir a consulta Oracle
    unidade = [
        (1, 'Centro de Tecnologia da Informação'),
        (2, 'Centro de Biotecnologia'),
        (3, 'Escola Superior de Tecnologia'),
        (4, 'Escola Normal Superior'),
    ]

    # Verifica se o usuário tem PADACES 708 (mockado)
    padaces_aceitos = None

    if request.user.is_authenticated:
        user_ldap_session = request.session.get('user_id', 'cn=fakeuser,ou=usuarios').split(',')[0].split('=')[1]

        # Simula que esse usuário tem o PADACES 708
        if user_ldap_session == 'fakeuser':
            padaces_aceitos = True

    is_admin = padaces_aceitos is not None

    if request.method == 'POST':
        try:
            unidade = request.POST.get('unidade')
            rua = request.POST.get('rua')
            numero_rua = request.POST.get('numero_rua')
            cep = request.POST.get('cep')
            bairro = request.POST.get('bairro')
            andar = request.POST.get('andar')
            sala = request.POST.get('sala')
            funcionamento = request.POST.get('funcionamento')

            laboratorio.unidade = unidade
            laboratorio.rua = rua
            laboratorio.numero_rua = numero_rua
            laboratorio.cep = cep
            laboratorio.bairro = bairro
            laboratorio.andar = andar
            laboratorio.sala = sala
            laboratorio.funcionamento = funcionamento

            laboratorio.save()
            messages.success(request, 'Endereço atualizado com sucesso.')

            return redirect('editar_endereco', laboratorio_id=laboratorio_id)
        except Exception as e:
            print('Erro:', str(e))
            messages.error(request, f'Erro ao atualizar o endereço, adicione apenas números.')

    return render(request, 'editar_endereco.html', {
        'laboratorio': laboratorio,
        'unidade': unidade,
        'is_admin': is_admin
    })

@login_required
@csrf_exempt
def visualizar_membros_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    membros = laboratorio.membros.all()

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        # Lógica para adicionar ou atualizar membros, se necessário
        nome_membro = request.POST.get('nome_membro')
        funcao = request.POST.get('funcao')
        curriculo_lattes = request.POST.get('curriculo_lattes')

        if nome_membro and funcao:
            membro, created = MembroLaboratorio.objects.get_or_create(
                laboratorio=laboratorio,
                nome_membro=nome_membro,
                defaults={'funcao': funcao, 'curriculo_lattes': curriculo_lattes}
            )

    return render(
        request,
        'visualizar_membros.html',
        {'laboratorio': laboratorio, 'membros': membros, 'is_admin': is_admin}
    )

@login_required
@csrf_exempt
def editar_membro_laboratorio(request, membro_id):
    membro = get_object_or_404(MembroLaboratorio, id=membro_id)
    laboratorio = membro.laboratorio

    # Verificação se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    # Permissão: só superusuário pode editar
    if not is_admin:
        return render(request, 'acesso_restrito.html', {
            'mensagem': 'Você não tem permissão para editar este membro.',
            'laboratorio': laboratorio,
        })

    if request.method == 'POST':
        nome_membro = request.POST.get('nome_membro')
        funcao = request.POST.get('funcao')
        curriculo_lattes = request.POST.get('curriculo_lattes')

        membro.nome_membro = nome_membro
        membro.funcao = funcao
        membro.curriculo_lattes = curriculo_lattes
        membro.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return render(
        request,
        'editar_membro_laboratorio.html',
        {'membro': membro, 'is_admin': is_admin}
    )
# novo
@login_required
@csrf_exempt
def excluir_membro_laboratorio(request, membro_id):
    membro = get_object_or_404(MembroLaboratorio, id=membro_id)
    membro.delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
@csrf_exempt
def adicionar_formento(request, laboratorio_id=None):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else request.user.laboratorio

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        # Restante do código para processar os campos do formulário e salvar as alterações
        nomedoparceito = request.POST.get('nomedoparceito')
        selecao = request.POST.get('selecao')
        edital = request.POST.get('edital')
        numerodoformento = request.POST.get('numerodoformento')
        tipo = request.POST.get('tipo')

        # Cria o objeto Formento
        formento = Formento.objects.create(
            nomedoparceito=nomedoparceito,
            selecao=selecao,
            tipo=tipo,
            edital=edital,
            numerodoformento=numerodoformento,
            laboratorio=laboratorio
        )

        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    formentos = laboratorio.Formentos.all()

    return render(
        request,
        'adicionar_formento.html',
        {'laboratorio': laboratorio, 'formentos': formentos, 'is_admin': is_admin}
    )

@login_required
@csrf_exempt
def adicionar_pos(request, laboratorio_id=None):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else request.user.laboratorio

    # Verifica se o usuário é superusuário do Django
    is_admin = request.user.is_superuser

    if request.method == 'POST':
        # Restante do código para processar os campos do formulário e salvar as alterações
        nome_do_Programa = request.POST.get('nome_do_Programa')
        website = request.POST.get('website')

        # Cria uma nova instância de Pos com os dados fornecidos
        pos = Pos.objects.create(
            nome_do_Programa=nome_do_Programa,
            website=website,
            laboratorio=laboratorio
        )

        # Redirecione para a página de edição do laboratório ou para onde for adequado
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    pos = laboratorio.pos.all()

    return render(
        request,
        'pos.html',
        {'laboratorio': laboratorio, 'pos': pos, 'is_admin': is_admin}
    )
@login_required
@csrf_exempt
def excluir_pos(request, pos_id):
    pos = get_object_or_404(Pos, id=pos_id)

    # Adicione a lógica para excluir o Pos
    if request.method == 'POST':
        pos.delete()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return render(
        request,
        'pos.html',
        {'pos': pos}
    )

@login_required
@csrf_exempt
def excluir_formento(request, formento_id):
    formento = get_object_or_404(Formento, id=formento_id)

    # Adicione a lógica para excluir o Formento
    if request.method == 'POST':
        formento.delete()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    return render(
        request,
        'excluir_formento.html',
        {'formento': formento}
    )

@login_required
@csrf_exempt
def editar_pos(request, laboratorio_id=None, pos_id=None):
    # Obtém o laboratório associado ao usuário ou ao laboratório_id fornecido
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else getattr(request.user, 'laboratorio', None)

    # Se formento_id não for fornecido, será uma adição
    if pos_id is None:
        return adicionar_pos(request, laboratorio_id)

    # Verifica se o objeto Formento existe
    pos = Pos.objects.filter(id=pos_id, laboratorio=laboratorio).first()
    if pos is None:
        raise Http404('A Pos não existe.')

    if request.method == 'POST':
        # Processa os dados diretamente da requisição POST
        nome_do_Programa = request.POST.get('nome_do_Programa')
        website = request.POST.get('website')

        # Atualiza os campos do objeto Formento
        pos.nome_do_Programa = nome_do_Programa
        pos.website = website
       
        pos.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    pos = laboratorio.pos.all()

    return render(
        request,
        'editar_pos.html',
        {'laboratorio': laboratorio, 'pos': pos}
    )

@login_required
@csrf_exempt
def editar_formento(request, laboratorio_id=None, formento_id=None):
    # Obtém o laboratório associado ao usuário ou ao laboratório_id fornecido
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id) if laboratorio_id else getattr(request.user, 'laboratorio', None)

    # Se formento_id não for fornecido, será uma adição
    if formento_id is None:
        return adicionar_formento(request, laboratorio_id)

    # Verifica se o objeto Formento existe
    formento = Formento.objects.filter(id=formento_id, laboratorio=laboratorio).first()
    if formento is None:
        raise Http404('O formento não existe.')

    if request.method == 'POST':
        # Processa os dados diretamente da requisição POST
        nomedoparceito = request.POST.get('nomedoparceito')
        selecao = request.POST.get('selecao')
        tipo = request.POST.get('tipo')
        
        edital = request.POST.get('edital')
        numerodoformento = request.POST.get('numerodoformento')

        # Atualiza os campos do objeto Formento
        formento.nomedoparceito = nomedoparceito
        formento.selecao = selecao
        formento.tipo = tipo
      
        formento.edital = edital
        formento.numerodoformento = numerodoformento
        formento.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    formentos = laboratorio.formentos.all()

    return render(
        request,
        'editar_formento.html',
        {'laboratorio': laboratorio, 'formentos': formentos}
    )


@csrf_exempt
def export_to_excel(request):

    if request.method == "POST":
        unidade = request.POST.getlist('unidade')
        export_all = request.POST.get('export_all')
        laboratorio = request.POST.getlist('laboratorios')
      
    
        if unidade:
            if len(unidade) == 1:
                laboratorios = Laboratorio.objects.filter(unidade=unidade[0])
           
                
            else:
                laboratorios = Laboratorio.objects.filter(unidade__in=unidade)
          
            equipamentos = Infraestrutura.objects.all()
            regimentos_internos = RegimentoInterno.objects.all()
            pos = Pos.objects.filter(laboratorio__in=laboratorios)
            formentos = Formento.objects.filter(laboratorio__in=laboratorios)
            projeto = Projeto.objects.filter(laboratorio__in=laboratorios)

        if export_all:
            laboratorios = Laboratorio.objects.all()
            equipamentos = Infraestrutura.objects.all()
            regimentos_internos = RegimentoInterno.objects.all()
            pos = Pos.objects.all()
            formentos = Formento.objects.all()
            projeto = Projeto.objects.all()

        if laboratorio:
            laboratorios = Laboratorio.objects.filter(id__in=laboratorio)
            equipamentos = Infraestrutura.objects.filter(laboratorio__in=laboratorios)
            regimentos_internos = RegimentoInterno.objects.filter(laboratorio__in=laboratorios)
            pos = Pos.objects.filter(laboratorio__in=laboratorios)
            formentos = Formento.objects.filter(laboratorio__in=laboratorios)
            projeto = Projeto.objects.filter(laboratorio__in=laboratorios)



    laboratorios_df = pd.DataFrame(laboratorios.values())
    
    if regimentos_internos.exists():
        regimentos_internos_df = pd.DataFrame(list(regimentos_internos.values()))
    else:
        regimentos_internos_df = pd.DataFrame()

    if pos.exists():
        pos_df = pd.DataFrame(list(pos.values()))
    else:
        pos_df = pd.DataFrame()

    if projeto.exists():
        projeto_df = pd.DataFrame(list(projeto.values()))
    else:
        projeto_df = pd.DataFrame()

    formentos_df = pd.DataFrame(list(formentos.values())) if formentos.exists() else pd.DataFrame() 
    equipamentos_df = pd.DataFrame(list(equipamentos.values()))
    
    # if 'laboratorio_id' in equipamentos_df.columns:
    #     equipamentos_df = equipamentos_df.drop(columns=['laboratorio_id'])


    equipamentos_df['marca_id'] = equipamentos_df['marca_id'].apply(lambda x: Marca.objects.get(id=x).nome_marca if (not isnan(x) and Marca.objects.filter(id=x).exists()) else 'Desconhecida')

    equipamentos_df['equipamento_id'] = equipamentos_df['equipamento_id'].apply(lambda x: Equipamento.objects.get(id=x).nome_equipamento if (not isnan(x) and Equipamento.objects.filter(id=x).exists()) else 'Desconhecida')    
    # combined_df = pd.merge(laboratorios_df, regimentos_internos_df, left_on='id', right_on='laboratorio_id', how='left')
    # combined_df = pd.merge(combined_df, equipamentos_df, left_on='laboratorio_id', right_on='id', how='left')
    
 

    combined_df = pd.merge(laboratorios_df, regimentos_internos_df, left_on='id', right_on='laboratorio_id', how='left')
    pos_grouped = pos_df.groupby('laboratorio_id').agg(list).reset_index()
    formentos_grouped = formentos_df.groupby('laboratorio_id').agg(list).reset_index()
    projeto_grouped = projeto_df.groupby('laboratorio_id').agg(list).reset_index()

    combined_df = pd.merge(combined_df, pos_grouped, left_on='laboratorio_id', right_on='laboratorio_id', how='left')
    combined_df = pd.merge(combined_df, formentos_grouped, left_on='laboratorio_id', right_on='laboratorio_id', how='left')
    combined_df = pd.merge(combined_df, projeto_grouped, left_on='laboratorio_id', right_on='laboratorio_id', how='left')

    
    equipamentos_grouped = equipamentos_df.groupby('laboratorio_id').agg(list).reset_index()
    combined_df = pd.merge(combined_df, equipamentos_grouped, left_on='laboratorio_id', right_on='laboratorio_id', how='left')
  

    combined_df['nomedoparceito'] = combined_df['nomedoparceito'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['selecao'] = combined_df['selecao'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
   
    combined_df['edital'] = combined_df['edital'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['numerodoformento'] = combined_df['numerodoformento'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['id_x'] = combined_df['id_x'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
  

    combined_df['nome_do_Programa'] = combined_df['nome_do_Programa'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['website'] = combined_df['website'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['equipamento_id'] = combined_df['equipamento_id'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['marca_id'] = combined_df['marca_id'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['modelo'] = combined_df['modelo'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['finalidade'] = combined_df['finalidade'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['status_y'] = combined_df['status_y'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['tombo'] = combined_df['tombo'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['quantidade'] = combined_df['quantidade'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['projeto'] = combined_df['projeto'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['nome_projeto'] = combined_df['nome_projeto'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['docente_responsavel'] = combined_df['docente_responsavel'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['discente_participante'] = combined_df['discente_participante'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['matricula_discente'] = combined_df['matricula_discente'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['modalidade'] = combined_df['modalidade'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['vigencia_inicio'] = combined_df['vigencia_inicio'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['vigencia_fim'] = combined_df['vigencia_fim'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    combined_df['fomento'] = combined_df['fomento'].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    
        
    combined_df = combined_df.drop(columns=[ 'id_x', 'id_y', 'discente_participante', 'matricula_discente'])
    combined_df = combined_df.drop_duplicates(subset=['nome_laboratorio'])  # Compara baseado no primeiro elemento

    print(combined_df)
    column_name_mapping = {
        'nome_laboratorio': 'LABORATÓRIO',
        'responsavel': 'RESPONSAVEL',
        'email': 'EMAIL',
        'telefone': 'TELEFONE',
        'unidade': 'UNIDADE',
        'rua': 'RUA',
        'numero_rua': 'NÚMERO',
        'cep': 'CEP',
        'bairro': 'BAIRRO',
        'andar': 'ANDAR',
        'sala': 'SALA',
        'apresentacao': 'APRESENTAÇÃO',
        'objetivos': 'OBJETIVOS',
        'descricao': 'DESCRICAO',
        'link_pnipe': 'LINK',
        'ato_anexo': 'ANEXO',
        'pdf': ' PDF',
        'cpf_responsavel': 'CPF',
        'funcionamento': 'FUNCIONAMENTO',
        'site': 'SITE',
        'nome_do_Programa': 'PROGRAMA',
        'nomedoparceito': 'NOME DO PARCEIRO',
        'selecao': 'SELEÇÃO',
       
        'edital': 'EDITAL',
        'numerodoformento': 'N° DO FORMENTO',
        'equipamento_id': 'EQUIPAMENTO',
        'marca_id': 'MARCA',
        'modelo': 'MODELO',
        'finalidade': 'FINALIDADE',
        'tombo': 'TOMBO',
        'quantidade': 'QUANTIDADE',
    }

    combined_df = combined_df.rename(columns=column_name_mapping)


    combined_df = combined_df.drop(columns=[ 'laboratorio_id',  'status_y', 'ANEXO', 'aprovado', 'user_ldap_responsavel', 'status_x', 'nome_do_pdf'])

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.title = 'Laboratório'

    fill_green_dark = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

    row_colors = ['00ff00', '4aea37', '70bf5d', '79aa6b', '7e9576', '808080']

    for col_idx, column_name in enumerate(combined_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_name)
        cell.fill = fill_green_dark

    for row_idx, row in enumerate(combined_df.itertuples(), 2):
        row_color = row_colors[row_idx % len(row_colors)]
        for col_idx, value in enumerate(row[1:], 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="laboratorios_data.xlsx"'

    workbook.save(response)

    return response


@login_required
@csrf_exempt
def delete_infra(request, laboratorio_id):

    Infraestrutura.objects.filter(id=laboratorio_id).delete()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@csrf_exempt
@login_required
def change_status(request, id):
    try:
        equipamento = Infraestrutura.objects.get(id=id)
    except Infraestrutura.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    equipamento.status = not equipamento.status
    equipamento.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
@csrf_exempt
def change_status_pdf(request, id):
    try:
        regimento = RegimentoInterno.objects.get(id=id)
    except Infraestrutura.DoesNotExist:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    regimento.status = not regimento.status
    regimento.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@csrf_exempt
def projetos(request, laboratorio_id):
   
    resultados = [
        ('PROJ001', 'Projeto de Pesquisa em Biotecnologia'),
        ('PROJ002', 'Estudo de Modelagem Computacional'),
        ('PROJ003', 'Projeto de Extensão em Comunidades Ribeirinhas'),
    ]

    # Verifica se o usuário tem PADACES 708 (simulado)
    padaces_aceitos = None

    if request.user.is_authenticated:
        user_ldap_session = request.session.get('user_id', 'cn=fakeuser,ou=usuarios').split(',')[0].split('=')[1]

        # Simula que esse usuário tem o PADACES 708
        if user_ldap_session == 'fakeuser':
            padaces_aceitos = True

    is_admin = padaces_aceitos is not None

    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    projetos = Projeto.objects.filter(laboratorio=laboratorio)

    if request.method == 'POST':
        form = ProjetoForm(request.POST)
        if form.is_valid():
            projeto = form.save(commit=False)
            projeto.laboratorio = laboratorio
            projeto.save()

            # Processar os discentes e matrículas
            discentes_input = request.POST.getlist('discentes')
            matriculas_input = request.POST.getlist('matriculas')

            for nome_discente, matricula in zip(discentes_input, matriculas_input):
                discente = Discente.objects.create(nome=nome_discente.strip(), projeto=projeto)
                Matricula.objects.create(discente=discente, matricula=matricula.strip())

            projetos = Projeto.objects.filter(laboratorio=laboratorio)

            return render(request, 'projetos.html', {
                'form': form,
                'projetos': projetos,
                'laboratorio': laboratorio,
                'resultados': resultados,
                'is_admin': is_admin
            })
    else:
        form = ProjetoForm(initial={'laboratorio': laboratorio_id, 'tem_cadastro_uea': True})

    return render(request, 'projetos.html', {
        'form': form,
        'projetos': projetos,
        'laboratorio': laboratorio,
        'resultados': resultados,
        'is_admin': is_admin
    })

@login_required
@csrf_exempt
def excluir_projeto(request, projeto_id):
    projeto = get_object_or_404(Projeto, id=projeto_id)
    laboratorio_id = projeto.laboratorio.id  # Obtém o laboratório associado ao projeto
    projeto.delete()
    return redirect('projetos', laboratorio_id=laboratorio_id)  # Redirecione de volta para a página de projetos após excluir

@csrf_exempt
def obter_nome_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')

    # Dados mockados simulando a tabela XPROJ2.PROJETO
    projetos_mock = {
        'PROJ001': 'Projeto de Pesquisa em Biotecnologia',
        'PROJ002': 'Estudo de Modelagem Computacional',
        'PROJ003': 'Projeto de Extensão em Comunidades Ribeirinhas',
    }

    nome_projeto = projetos_mock.get(projeto_selecionado, '')

    return JsonResponse({'nome_projeto': nome_projeto})


@csrf_exempt
def obter_detalhes_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')

    # Dados mockados para simular a base XPROJ2
    projetos_mock = {
        'PROJ001': {
            'titulo': 'Projeto de Pesquisa em Biotecnologia',
            'tipo': 'T01',
            'autor': 'iury.uea'
        },
    }

    membros_mock = {
        'PROJ001': [
            '12345678901',  # CPF válido (para EXTERNO)
            'MATRIC9876',   # Matrícula não CPF
        ]
    }

    externo_mock = {
        '12345678901': 'João da Silva',
    }

    tipo_mock = {
        'T01': 'Pesquisa Científica'
    }

    fomento_mock = {
        'PROJ001': 'CNPq'
    }

    # Simulando a base OBERON
    usuario_mock = {
        'iury.uea': '12345678901'
    }

    pessoa_mock = {
        '12345678901': 'Iury Costa de Souza'
    }

    # Obter dados do projeto
    projeto = projetos_mock.get(projeto_selecionado)
    if not projeto:
        return JsonResponse({'error': 'Projeto não encontrado'}, status=404)

    nome_projeto = projeto['titulo']
    nome_autor_projeto = projeto['autor']
    tipo_projeto = projeto['tipo']

    # Buscar modalidade
    modalidade_projeto = tipo_mock.get(tipo_projeto, '')

    # Buscar discentes
    nomes_discentes = []
    matriculas = membros_mock.get(projeto_selecionado, [])
    for matricula in matriculas:
        if len(matricula) == 11 and matricula.isdigit():  # Simulando CPF
            nome_externo = externo_mock.get(matricula)
            nomes_discentes.append(nome_externo if nome_externo else matricula)
        else:
            nomes_discentes.append(matricula)

    matricula_discente = ', '.join(matriculas)

    # Buscar fomento
    fomento_projeto = fomento_mock.get(projeto_selecionado, 'Fomento não encontrado no banco de dados.')

    # Buscar dados do autor (OBERON)
    cpf_autor = usuario_mock.get(nome_autor_projeto, 'CPF não encontrado na tabela usuario.')
    nome_completo_autor = pessoa_mock.get(cpf_autor, 'Nome não encontrado na tabela pessoa.')

    return JsonResponse({
        'nome_projeto': nome_projeto,
        'nome_autor_projeto': nome_autor_projeto,
        'nomes_discentes': nomes_discentes,
        'modalidade_projeto': modalidade_projeto,
        'matricula_discente': matricula_discente,
        'fomento_projeto': fomento_projeto,
        'cpf_autor': cpf_autor,
        'nome_completo_autor': nome_completo_autor
    })

@csrf_exempt
def obter_vigencia_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')

    if not projeto_selecionado:
        return JsonResponse({'error': 'Projeto não fornecido.'}, status=400)

    # Simulação (mock) de dados para teste
    projetos_mock = {
        'P001': {'vigencia_inicio': '2023-01-01', 'vigencia_fim': '2023-12-31'},
        'P002': {'vigencia_inicio': '2022-05-10', 'vigencia_fim': '2023-05-09'},
        'P003': {'vigencia_inicio': '2024-03-01', 'vigencia_fim': '2025-02-28'},
    }

    dados_vigencia = projetos_mock.get(projeto_selecionado)

    if not dados_vigencia:
        return JsonResponse({'error': 'Projeto não encontrado.'}, status=404)

    return JsonResponse({
        'vigencia_inicio': dados_vigencia['vigencia_inicio'],
        'vigencia_fim': dados_vigencia['vigencia_fim']
    })


@csrf_exempt
def obter_modalidade_projeto(request):
    projeto_selecionado = request.GET.get('projeto', '')

    if not projeto_selecionado:
        return JsonResponse({'error': 'Projeto não fornecido.'}, status=400)

    # Dados simulados (mockados)
    projetos_mock = {
        'P001': 'T01',
        'P002': 'T02',
        'P003': 'T03',
    }

    tipos_mock = {
        'T01': 'Pesquisa Científica',
        'T02': 'Desenvolvimento Tecnológico',
        'T03': 'Inovação Acadêmica',
    }

    tipo_projeto = projetos_mock.get(projeto_selecionado)

    if not tipo_projeto:
        return JsonResponse({'error': 'Tipo do projeto não encontrado.'}, status=404)

    modalidade = tipos_mock.get(tipo_projeto, '')

    return JsonResponse({'modalidade': modalidade})

@login_required
@csrf_exempt
def aprovar_laboratorio(request, laboratorio_id):
    # Verifica se o usuário é superusuário
    if not request.user.is_superuser:
        raise PermissionDenied("Você não tem permissão para aprovar laboratórios.")

    # Lógica para aprovar o laboratório
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)
    laboratorio.aprovado = True
    laboratorio.save()

    messages.success(request, 'Laboratório aprovado com sucesso.')
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)

@login_required
@csrf_exempt
def cancelar_aprovacao(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Adicione lógica para cancelar a aprovação aqui, por exemplo, atualizando o campo 'aprovado'
    laboratorio.aprovado = False
    laboratorio.save()

    messages.success(request, 'Aprovação do laboratório cancelada com sucesso.')
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)

@csrf_exempt
def processar_impressao(request):
    # Aqui você pode adicionar o código para processar a impressão
    # Por enquanto, deixarei a view simplesmente retornando a página de visualização
    return render(request, 'view.html')  # Você precisa criar o template processar_impressao.html


@csrf_exempt
@login_required
@requer_padace(708)
def ocultar_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Adicione a lógica para ocultar o laboratório aqui, por exemplo, definindo o campo 'ocultar_laboratorio' como True
    laboratorio.ocultar_laboratorio = True
    laboratorio.save()

    messages.success(request, 'Laboratório ocultado com sucesso.')
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)


@csrf_exempt
@login_required
def cancelar_ocultacao(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Adicione a lógica para cancelar a ocultação do laboratório aqui
    laboratorio.ocultar_laboratorio = False
    laboratorio.save()

    messages.success(request, 'Ocultação do laboratório cancelada com sucesso.')
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)


@csrf_exempt
@login_required
@requer_padace(708)
def desocultar_laboratorio(request, laboratorio_id):
    laboratorio = get_object_or_404(Laboratorio, id=laboratorio_id)

    # Lógica para desocultar o laboratório
    laboratorio.ocultar_laboratorio = False
    laboratorio.save()

    # Mensagem de sucesso (opcional)
    messages.success(request, 'Laboratório desocultado com sucesso.')

    # Redirecionamento para a página desejada após desocultar o laboratório
    return redirect('editar_laboratorio', laboratorio_id=laboratorio.id, imagem_id=0)


from django.shortcuts import render, get_object_or_404, redirect
from .forms import ProjetoForm, ParticipantesForm
from .models import Projeto, Discente, Matricula

@csrf_exempt
@login_required
def editar_projeto(request, projeto_id):
    projeto = get_object_or_404(Projeto, id=projeto_id)

    if request.method == 'POST':
        projeto_form = ProjetoForm(request.POST, instance=projeto)
        participantes_form = ParticipantesForm(request.POST)
        if projeto_form.is_valid() and participantes_form.is_valid():
            projeto_form.save()

            # Processamento dos discentes e matrículas
            discentes_input = request.POST.getlist('discentes')
            matriculas_input = request.POST.getlist('matriculas')

            # Limpa os discentes existentes associados a este projeto
            projeto.discentes.all().delete()

            # Atualiza ou cria os discentes e matrículas
            for nome_discente, matricula in zip(discentes_input, matriculas_input):
                if nome_discente.strip():
                    discente = Discente.objects.create(nome=nome_discente.strip(), projeto=projeto)
                    Matricula.objects.create(discente=discente, matricula=matricula.strip())

            return redirect('projetos', laboratorio_id=projeto.laboratorio.id)
    else:
        projeto_form = ProjetoForm(instance=projeto)
        discentes = projeto.get_discentes_list()
        matriculas = projeto.get_matriculas_list()
        participantes_data = zip(discentes, matriculas)
        participantes_form = ParticipantesForm(initial={'discentes': '\n'.join(discentes), 'matriculas': '\n'.join(matriculas)})

    return render(request, 'editar_projeto.html', {'projeto_form': projeto_form, 'participantes_form': participantes_form, 'projeto': projeto, 'participantes_data': participantes_data})

    
    
    
@csrf_exempt
def get_infra_data(request):
    infra_id = request.GET.get('infra_id')  # Obtém o ID da infraestrutura a partir da solicitação AJAX
    infra = Infraestrutura.objects.get(id=infra_id)

    # Crie um dicionário com os dados que deseja retornar
    infra_data = {
        'modelo': infra.modelo,
        'finalidade': infra.finalidade,
        'tombo': infra.tombo,
        'quantidade': infra.quantidade,
        # Adicione outros campos conforme necessário
    }

    return JsonResponse(infra_data)
    
    
from django.http import JsonResponse
from .models import Infraestrutura

@csrf_exempt
@login_required
def salvar_edicao_infra(request):
    if request.method == 'POST':
        infra_id = request.POST.get('infra_id')
        modelo = request.POST.get('modelo')
        finalidade = request.POST.get('finalidade')
        tombo = request.POST.get('tombo')
        quantidade = request.POST.get('quantidade')
        equipamento_id = request.POST.get('edit_equipamento')
        marca_id = request.POST.get('edit_marca')

        infra = Infraestrutura.objects.get(id=infra_id)
        infra.modelo = modelo
        infra.finalidade = finalidade
        infra.tombo = tombo
        infra.quantidade = quantidade
        infra.equipamento_id = equipamento_id
        infra.marca_id = marca_id

        # Verifique se novos arquivos de imagem foram enviados
        novas_imagens = request.FILES.getlist('edit_imagem')

        # Adicione as novas imagens à infraestrutura existente
        for nova_imagem in novas_imagens:
            nova_imagem_instance = ImagemInfraestrutura(imagem=nova_imagem, infraestrutura=infra)
            nova_imagem_instance.save()

        infra.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Método não permitido'})


from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import ProjetoForm, DiscenteForm
from .models import Projeto, Discente, Matricula

from .models import Projeto, Discente, Matricula

@csrf_exempt
@login_required
def criar_projeto(request, laboratorio_id):
    if request.method == 'POST':
        # Adicione o laboratório ao formulário antes de validar
        mutable_post = request.POST.copy()
        mutable_post['laboratorio'] = laboratorio_id
        projeto_form = ProjetoForm(mutable_post)
        discente_form = DiscenteForm(request.POST)

        if projeto_form.is_valid() and discente_form.is_valid():
            projeto = projeto_form.save(commit=False)
            projeto.save()

            # Processamento dos discentes e matrículas
            discentes_input = request.POST.get('discentes', '').split('\n')
            matriculas_input = request.POST.get('matriculas', '').split('\n')

            for nome_discente, matricula in zip(discentes_input, matriculas_input):
                discente = Discente.objects.create(nome=nome_discente.strip(), projeto=projeto)
                Matricula.objects.create(discente=discente, matricula=matricula.strip())

            return redirect('projetos', laboratorio_id=laboratorio_id)  # Redirecionar para a página de projetos
    else:
        # Se for uma solicitação GET, apenas crie os formulários vazios
        projeto_form = ProjetoForm(initial={'laboratorio': laboratorio_id})  # Defina o laboratório como valor inicial
        discente_form = DiscenteForm()

    return render(request, 'projetos.html', {'projeto_form': projeto_form, 'discente_form': discente_form})

